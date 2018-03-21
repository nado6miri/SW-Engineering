import sys
import os
import copy

#for Jira control
from jira import JIRA
#for excel control
import xlsxwriter as xlswt
import openpyxl as xlsrd
#for time
from datetime import datetime
# for UI
from PyQt5 import uic, QtWidgets, QtGui

#http://hlm.lge.com/issue/rest/api/2/issue/GSWDIM-22476/
#http://hlm.lge.com/issue/rest/api/2/issue/TVPLAT-3963/

#http://hlm.lge.com/qi/rest/api/2/issue/QEVENTSEVT-7232/ - Q


DevTracker = 'http://hlm.lge.com/issue'
QTracker = 'http://hlm.lge.com/qi'

userID = 'sungbin.na'
userPasswd = 'Sungbin@1801'

'''
dissue_dict = {
    'project': {'key': 'GSWDIM'},
    'components' : [ ],
    'summary': 'New issue from jira-python',
    'description': 'Look into this one',
    'parent' : { 'id' :  ''},
    'issuetype' : { 'name' : 'Sub-task' },
    #'issuetype': {'id': '5'},
    'assignee': {"name":"sungbin.na", "emailAddress":"sungbin.na@lge.com"},
    'reporter': {"name":"sungbin.na", "emailAddress":"sungbin.na@lge.com"},
    'labels' : ['Default_label'],
    'duedate' : '2018-04-30',
    #'customfield_10105' :[{"name":"sungbin.na","key":"sungbin.na","emailAddress":"sungbin.na@lge.com" },] #watchers
    'customfield_10105' :[ {"name":"sungbin.na" }, {"name":"insun.song" }], #watchers
    'comment' : { 'comments' : [ { 'body' : ''}, ] }, #comment
}
'''


dissue_dict = {
'''
    'project': {'key': ''},
    'components' : [ ],
    'summary': '',
    'description': '',
    'parent' : { 'id' :  ''},
    'issuetype' : { 'name' : '' },
    #'issuetype': {'id': '5'},
    'assignee': { },
    'reporter': { },
    'labels' : [ ],
    'duedate' : '',
    #'customfield_10105' :[{"name":"sungbin.na","key":"sungbin.na","emailAddress":"sungbin.na@lge.com" },] #watchers
    'customfield_10105' :[ ], #watchers
    'comment' : { 'comments' : [ { 'body' : ''}, ] }, #comment
    '''
}

dissue_init_dict = {
    'project': {'key': ''},
    'components' : [ ],
    'summary': '',
    'description': '',
    'parent' : { 'id' :  ''},
    'issuetype' : { 'name' : '' },
    #'issuetype': {'id': '5'},
    'assignee': { },
    'reporter': { },
    'labels' : [ ],
    'duedate' : '',
    #'customfield_10105' :[{"name":"sungbin.na","key":"sungbin.na","emailAddress":"sungbin.na@lge.com" },] #watchers
    'customfield_10105' :[ ], #watchers
    #'comment' : { 'comments' : [ { 'body' : ''}, ] }, #comment
}

def makeKeyList() :
    # read header and make key list to make jira json file
    keylist = []
    j = 1
    cols = ws.columns
    for col in cols :
        val = ws.cell(row = 1, column = j).value
        if(val is not None):
            keylist.append(val)
            j+=1
        else :
            pass
    return keylist

def setProject(keyword, value) :
    if(value is not None):
        print(keyword, " = ", value)
        dissue_dict[keyword]['key'] = value
    else :
        print(keyword, " = None... Skip")


def setComponents(keyword, value) :
    if(value is not None):
        print(keyword, " = ", value)
        comp_list = value.split(',')
        print(comp_list)
        comp_dict = { 'name' : ''}
        for cl in comp_list :
            comp_dict['name'] = cl.strip()
            dissue_dict[keyword].append(comp_dict)
            #print(comp)
        #print(dissue_dict[keyword])
    else :
        print(keyword, " = None... Skip")


def setIssueType(keyword, value) :
    if(value is not None):
        print(keyword, " = ", value)
        dissue_dict[keyword]['name'] = value.strip()
    else :
        print(keyword, " = None... Skip")

def setParent(keyword, value) :
    if(value is not None):
        print(keyword, " = ", value)
        dissue_dict[keyword]['id'] = value
    else :
        del dissue_dict[keyword]
        print(keyword, " = None... Skip")

def setSummarynDescription(keyword, value) :
    if(value is not None):
        print(keyword, " = ", value)
        dissue_dict[keyword] = value.strip()
    else :
        print(keyword, " = None... Skip")

def setAssigneenReporter(keyword, value) :
    if(value is not None):
        print(keyword, " = ", value)
        dissue_dict[keyword]['name'] = value.strip()
    else :
        print(keyword, " = None... Skip")

def setWatchers(keyword, value) :
    if(value is not None):
        print(keyword, " = ", value)
        watcher_list = value.split(',')
        #print(watcher_list)
        for watcher in watcher_list :
            watcher_dict = { 'name' : ''}
            watcher_dict['name'] = watcher.strip() # delete space
            dissue_dict['customfield_10105'].append(watcher_dict)
            print("========================")
            print(watcher.strip())
    else :
        print(keyword, " = None... Skip")

def setDuedate(keyword, value) :
    if(value is not None):
        #duedate = datetime.strptime(value, '%Y-%m-%d')
        print(keyword, " = ", value)
        dissue_dict[keyword] = str(value)
    else :
        print(keyword, " = None... Skip")

def setLabels(keyword, value) :
    if(value is not None):
        label_list = value.split(',')
        print(label_list)
        for label in label_list :
            dissue_dict[keyword].append(label)
    else :
        print(keyword, " = None... Skip")

def setComment(keyword, value) :
    if(value is not None):
        print(keyword, " = ", value)
        #dissue_dict['comment']['comments']['0']['body'] = value
        #'comment' : { 'comments' : [ { 'body' : ''}, ] }, #comment
    else :
        print(keyword, " = None... Skip")

def setAttachment(keyword, value) :
    if(value is not None):
        print(keyword, " = ", value)
    else :
        print(keyword, " = None... Skip")

def setCommonNotice(keyword, value) :
    if(value is not None):
        print(keyword, " = ", value)
        desc = dissue_dict['description']
        desc = str(desc) + str(value)
        dissue_dict['description'] = desc
        #print("========================")
        #print(dissue_dict['description'])
        #print("========================")
    else :
        print(keyword, " = None... Skip")



def makeDevJiraJSON(key, value) :
    if (key == 'project') :
        setProject(key, value)
    elif (key == 'components'):
        setComponents(key, value)
    elif (key == 'issuetype'):
        print(key + ' = ' + value)
        setIssueType(key, value)
    elif (key == 'parent'):
        setParent(key, value)
    elif (key == 'summary'):
        setSummarynDescription(key, value)
    elif (key == 'description'):
        setSummarynDescription(key, value)
    elif (key == 'assignee'):
        setAssigneenReporter(key, value)
    elif (key == 'reporter'):
        setAssigneenReporter(key, value)
    elif (key == 'watcher'):
        setWatchers(key, value)
    elif (key == 'duedate'):
        setDuedate(key, value)
    elif (key == 'labels'):
        setLabels(key, value)
    elif (key == 'comment'):
        setComment(key, value)
    elif (key == 'attachment'):
        print("Set attachment")
    elif (key == 'Common Notice'):
        setCommonNotice(key, value)
    else :
        print("[Error] Set default="+key)



if __name__ == "__main__" :
    # 엑셀파일 열기
    excel_file = xlsrd.load_workbook('Jira_Issue.xlsm')
    ws = excel_file['Dev Tracker']

    jira_keylist = makeKeyList()
    print(jira_keylist)

    dev_jira = JIRA(DevTracker, basic_auth = (userID, userPasswd))
    q_jira = JIRA(QTracker, basic_auth = (userID, userPasswd))


    i = 1; j = 1
    rows = ws.rows
    for row in rows :
        if(i == 1) : i+=1; j = 1; continue
        dissue_dict = copy.deepcopy(dissue_init_dict)
        for key in jira_keylist :
            if(key == 'Key') :
                j = 2; continue
            if(key == 'Common Notice') :
                val = ws.cell(row = 2, column = j).value
            else :
                val = ws.cell(row = i, column = j).value
            #print("=====================")
            #print(key, val)
            #print("=====================")
            makeDevJiraJSON(key, val)
            j += 1
        try :
            print("========================================================")
            print(dissue_dict)
            new_dissue = dev_jira.create_issue(fields=dissue_dict)
            createdkey = new_dissue.raw['key']
            print("Created Key = ", createdkey)
            ws.cell(row = i, column = 1).value = createdkey
            print("========================================================")
        except Exception as e:
            print(e)
        i += 1

        #excel_file.save('Jira_Issue.xlsm')




    '''
    # 모든 row 접근
    rows = ws.rows
    for row in rows :
        print(row[0].value)
        print(row[1].value)
        print(row[2].value)

    # 모든 col 접근
    cols = ws.columns
    for col in cols :
        print(col[0].value)
        print(col[1].value)
        print(col[2].value)
    '''

    '''
    dev_jira = JIRA(DevTracker, basic_auth = (userID, userPasswd))
    q_jira = JIRA(QTracker, basic_auth = (userID, userPasswd))

    #====================================================================================================================
    # Dev Trakcer
    #====================================================================================================================
    # Get issue with All Fields in Dev Tracker
    dissue = dev_jira.issue("LEADSWETDI-1")
    print("[Dev Tracker] Get JIRA Issue with All fields")


    #====================================================================================================================
    # How to create Jira Issue in Dev Tracker
    #====================================================================================================================
    # Case 1:
    new_dissue = dev_jira.create_issue(project='GSWDIM', summary='New issue from jira-python', description='Look into this one', issuetype={'name': 'Bug'})

    # Case 2:
    issue_dict = {
        'project': {'key': 'GSWDIM'},
        'components' : [ { 'name' : 'JIRATEST' } ],
        'summary': 'New issue from jira-python',
        'description': 'Look into this one',
        'issuetype': {'name': 'Bug'},
        'assignee': {"name":"sungbin.na", "emailAddress":"sungbin.na@lge.com"},
        'labels' : ['Default_label'],
        'duedate' : '2018-04-30',
        #'customfield_10105' :[{"name":"sungbin.na","key":"sungbin.na","emailAddress":"sungbin.na@lge.com" },] #watchers
        'customfield_10105' :[ {"name":"sungbin.na" }, {"name":"insun.song" }] #watchers

    }
    issue_dict['labels'].append("VTASK")
    new_dissue = dev_jira.create_issue(fields=issue_dict)

    # Case 3:
    issue_list = [
    {
        'project': {'key': 'GSWDIM'},
        'components' : [ { 'name' : 'JIRATEST' } ],
        'summary': 'New issue from jira-python1',
        'description': 'Look into this one',
        'issuetype': {'name': 'Bug'},
        'assignee': {"name":"sungbin.na", "emailAddress":"sungbin.na@lge.com"},
        'labels' : ['Default_label'],
        'duedate' : '2018-04-30',
        #'customfield_10105' :[{"name":"sungbin.na","key":"sungbin.na","emailAddress":"sungbin.na@lge.com" },] #watchers
        'customfield_10105' :[ {"name":"sungbin.na" }, ] #watchers
    },
    {
        'project': {'key': 'GSWDIM'},
        'components' : [ { 'name' : 'JIRATEST' } ],
        'summary': 'New issue from jira-python2',
        'description': 'Look into this one',
        'issuetype': {'name': 'Bug'},
        'assignee': {"name":"sungbin.na", "emailAddress":"sungbin.na@lge.com"},
        'labels' : ['Default_label'],
        'duedate' : '2018-04-30',
        #'customfield_10105' :[{"name":"sungbin.na","key":"sungbin.na","emailAddress":"sungbin.na@lge.com" },] #watchers
        'customfield_10105' :[ {"name":"sungbin.na" }, ] #watchers
    },
    {
        'project': {'key': 'GSWDIM'},
        'components' : [ { 'name' : 'JIRATEST' } ],
        'summary': 'New issue from jira-python3',
        'description': 'Look into this one',
        'issuetype': {'name': 'Bug'},
        'assignee': {"name":"sungbin.na", "emailAddress":"sungbin.na@lge.com"},
        'labels' : ['Default_label'],
        'duedate' : '2018-04-30',
        #'customfield_10105' :[{"name":"sungbin.na","key":"sungbin.na","emailAddress":"sungbin.na@lge.com" },] #watchers
        'customfield_10105' :[ {"name":"sungbin.na" }, ] #watchers
    },
    ]
    new_dissue = dev_jira.create_issue(fields=issue_list)

    #====================================================================================================================
    # How to create SubTask in Dev Tracker
    #====================================================================================================================
    subissue_dict = {
        'project': {'key': 'GSWDIM'},
        'components' : [ { 'name' : 'JIRATEST' } ],
        'summary': 'New issue from jira-python',
        'description': 'Look into this one',
        'parent' : { 'id' :  ''},
        'issuetype' : { 'name' : 'Sub-task' },
        #'issuetype': {'id': '5'},
        'assignee': {"name":"sungbin.na", "emailAddress":"sungbin.na@lge.com"},
        'labels' : ['Default_label'],
        'duedate' : '2018-04-30',
        #'customfield_10105' :[{"name":"sungbin.na","key":"sungbin.na","emailAddress":"sungbin.na@lge.com" },] #watchers
        'customfield_10105' :[ {"name":"sungbin.na" }, {"name":"insun.song" }] #watchers
    }

    subissue_dict['parent']['id'] = new_dissue.key
    subissue_dict['labels'].append("VTASK")
    new_dissue = dev_jira.create_issue(fields=subissue_dict)


    #====================================================================================================================
    # How to update Jira Info in Dev Tracker
    #====================================================================================================================
    # Case 1:
    dissue = dev_jira.issue("LEADSWETDI-1")
    dissue.update(notify=True, summary='new summary', description='A new summary was added')

    # Case 2:
    dissue = dev_jira.issue("GSWDIM-22479")
    updateissue_dict = {
        'components' : [ { 'name' : 'JIRATEST' } ],
        'summary': 'New issue from jira-python - Update',
        'description': 'Look into this one - Update',
        'assignee': {"name":"insun.song"},
        'labels' : ['Default_label_Update'],
        'duedate' : '2018-05-30',
        #'customfield_10105' :[{"name":"sungbin.na","key":"sungbin.na","emailAddress":"sungbin.na@lge.com" },] #watchers
        'customfield_10105' :[{"name":"insun.song" }] #watchers

    }
    dissue.update(notify=True, fields = updateissue_dict)
    #issue.update(notify=True, assignee={'name': 'insun.song'})
    dissue.update(labels=['AAA', 'BBB'])


    #====================================================================================================================
    # How to reassign an Issue in Dev Tracker
    #====================================================================================================================
    issue = dev_jira.issue("GSWDIM-22479")
    dev_jira.assign_issue(issue, 'sungbin.na')


    #====================================================================================================================
    # How to delete an Issue in Dev Tracker
    #====================================================================================================================
    issue = dev_jira.issue("GSWDIM-22475") # if the issue has subtasks, can't delete it.
    issue.delete()


    #====================================================================================================================
    # How to add comment to Issue in Dev Tracker
    #====================================================================================================================
    # Case 1:
    comment = dev_jira.add_comment('GSWDIM-22479', 'new comment')    # no Issue object required

    # Case 2:
    issue = dev_jira.issue("GSWDIM-22479")
    comment = dev_jira.add_comment(issue, 'new comment', visibility={'type': 'role', 'value': 'Administrators'})  # for admins only
    comment.update(body = 'updated comment body')
    comment.delete()

    # Case 3:
    dissue = dev_jira.issue("GSWDIM-22479")
    comment = dissue.raw['fields']['comment']
    comments = comment['comments']
    for c in comments :
        print(c)
        print("===================================")
        print(c['body'])
        print("===================================")


    #====================================================================================================================
    # How to Add / Remove Watcher to Issue in Dev Tracker
    #====================================================================================================================
    issue = dev_jira.issue("GSWDIM-22479")
    watcher = dev_jira.watchers(issue)
    print("Issue has {} watcher(s)".format(watcher.watchCount))
    for watcher in watcher.watchers:
        print(watcher)
        # watcher is instance of jira.resources.User:
        print(watcher.emailAddress)

    dev_jira.add_watcher(issue, 'sungbin.na')
    dev_jira.revmove_watcher(issue, 'sungbin.na')



    #===========================================================================
    # Attachment control in Dev Tracker
    issue = dev_jira.issue("GSWDIM-22479")
    #jira.add_attachment(issue=issue, attachment='Jira_자동_등록 - webOS4.0 Issue.xlsm')

    # read and upload a file (note binary mode for opening, it's important):
    with open('Jira_자동_등록 - webOS4.0 Issue.xlsm', 'rb') as f:
        jira.add_attachment(issue=issue, attachment=f)

    # attach file from memory (you can skip IO operations). In this case you MUST provide `filename`.
    import StringIO
    attachment = StringIO.StringIO()
    attachment.write(data)
    jira.add_attachment(issue=issue, attachment=attachment, filename='content.txt')


    #===========================================================================
    # Q Trakcer
    #===========================================================================
    # Get issue with All Fields in Q Tracker
    qissue = q_jira.issue("WOSLQEVENT-98853")
    print("[Dev Tracker] Get JIRA Issue with All fields")

    # Get issue with specific Fields in Q Tracker
    setfield = ('summary, comment, assignee')
    qissue = q_jira.issue("LEADSWETDI-1", fields=setfield)
    print("[Dev Tracker] Get JIRA Issue with Specific fields")

    #===========================================================================
    # Get filtered issue with Filter ID in Q Tracker
    setFilterID = 'filter = 95949'
    qissue = q_jira.search_issues(setFilterID)
    print("[QTracker] Get JIRA Issue with Specific Filter ID: " + setFilterID)

    # Get Filtered issue with JQL Querfy String in Q Tracker
    setFilter = 'Filter in (M3.LK61.EU.QA1, M3.LK61.EU.QA2, M3.LK61.EU.QA3, M3.LK61.EU.QA4)'
    qissue = q_jira.search_issues(setFilter)
    print("[QTracker] Get JIRA Issue with Specific Filter String: " + setFilter)
    '''
