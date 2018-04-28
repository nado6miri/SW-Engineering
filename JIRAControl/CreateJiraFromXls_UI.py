import sys
import os
import copy
import time

#for Jira control
from jira import JIRA
from jira.exceptions import JIRAError

#for excel control
import xlsxwriter as xlswt
import openpyxl as xlsrd
#for time
from datetime import datetime
# for UI
#from PyQt5 import uic, QtWidgets, QtGui
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5 import uic

# dir(jira)
# dir(jira.JIRA)
# hasattr(JIRA, crete_issue)


#http://hlm.lge.com/issue/rest/api/2/issue/GSWDIM-22476/

#http://hlm.lge.com/issue/rest/api/2/issue/TVPLAT-3963/
#http://hlm.lge.com/issue/rest/api/2/issue/TVPLAT-3963/editmeta

#http://hlm.lge.com/qi/rest/api/2/issue/QEVENTSEVT-7232/ - Q


DevTracker = 'http://hlm.lge.com/issue'
QTracker = 'http://hlm.lge.com/qi'


# Excel
excel_file = None
wsheet = None
# Jira
dev_jira = False
q_jira = False

dissue_dict = {}
dissue_init_dict = {
    'project': {'key': ''},
    'components' : [ ],
    'summary': '',
    'description': '',
    'parent' : { 'id' :  ''},
    'issuetype' : { 'name' : '' },
    #'issuetype': {'id': '5'},
    'assignee': { },
    'reporter': { },
    'labels' : [ ],
    'duedate' : '',
    #'customfield_10105' :[{"name":"sungbin.na","key":"sungbin.na","emailAddress":"sungbin.na@lge.com" },] #watchers
    'customfield_10105' :[ ], #watchers
    'customfield_10436':'', # Epic Name
    #'comment' : { 'comments' : [ { 'body' : ''}, ] }, #comment
}

Islogin = False
Isexcelloaded = False


def makeKeyList(ws) :
    # read header and make key list to make jira json file
    keylist = []
    j = 1
    cols = ws.columns
    for col in cols :
        val = ws.cell(row = 1, column = j).value
        if(val is not None):
            keylist.append(val)
            j+=1
        else :
            pass
    return keylist

def setProject(keyword, value) :
    if(value is not None):
        print(keyword, " = ", value)
        dissue_dict[keyword]['key'] = value
    else :
        print(keyword, " = None... Skip")


def setComponents(keyword, value) :
    if(value is not None):
        value = str(value)
        print(keyword, " = ", value)
        comp_list = value.split(',')
        print(comp_list)
        comp_dict = { 'name' : ''}
        for cl in comp_list :
            comp_dict['name'] = cl.strip()
            dissue_dict[keyword].append(comp_dict)
            #print(comp)
        #print(dissue_dict[keyword])
    else :
        del dissue_dict[keyword]
        print(keyword, " = None... Skip")


def setIssueType(keyword, value) :
    if(value is not None):
        print(keyword, " = ", value)
        dissue_dict[keyword]['name'] = value.strip()
    else :
        dissue_dict[keyword]['name'] = 'Task'
        print(keyword, " = None... Set Default - Task")

def setParent(keyword, value) :
    if(value is not None):
        if(dissue_dict['issuetype']['name'] == 'Sub-task') :
            print(keyword, " = ", value)
            dissue_dict[keyword]['id'] = value
        else :
            del dissue_dict[keyword]
            print("Issue type is not Sub-Task.... No need to set parent id... delete this keyword")
    else :
        if(dissue_dict['issuetype']['name'] == 'Sub-Task') :
            print("[Err] Issue type is Sub-Task.... Need to set parent id...")
        else :
            del dissue_dict[keyword]
            print(keyword, " = Issue type is not Sub-Task, No need to set parent id... delete this keyword")

def setEpicName(keyword, value) :
    if(value is not None):
        print("************************")
        print(keyword, " = ", value)
        if(dissue_dict['issuetype']['name'] == 'Epic') :
            dissue_dict['customfield_10436'] = value
        else :
            del dissue_dict['customfield_10436']
            print("Issue type is not Epic.... No need to set Epic Name... delete this keyword")
    else :
        if(dissue_dict['issuetype']['name'] == 'Epic') :
            print("[Err] Issue type is Epic.... Need to set Epic Name...")
        else :
            del dissue_dict['customfield_10436']
            print(keyword, " = Issue type is not Epic, No need to set Epic Name... delete this keyword")


def setSummarynDescription(keyword, value) :
    if(value is not None):
        print(keyword, " = ", value)
        dissue_dict[keyword] = value.strip()
    else :
        if(keyword == 'summary') :
            dissue_dict[keyword] = 'Set Default Summary - Please change the summary later'
        elif (keyword == 'description') :
            dissue_dict[keyword] = 'Set Default Description - Please change the Description later'
        else :
            print(keyword, " = None... Skip")

def setAssigneenReporter(keyword, value) :
    if(value is not None):
        print(keyword, " = ", value)
        dissue_dict[keyword]['name'] = value.strip()
    else :
        if(keyword == 'assignee') :
            dissue_dict[keyword]['name'] = None
        elif (keyword == 'reporter') :
            dissue_dict[keyword]['name'] = 'hlm-admin'
        else :
            print(keyword, " = None... Skip")

def setWatchers(keyword, value) :
    if(value is not None):
        print(keyword, " = ", value)
        watcher_list = value.split(',')
        #print(watcher_list)
        for watcher in watcher_list :
            watcher_dict = { 'name' : ''}
            watcher_dict['name'] = watcher.strip() # delete space
            dissue_dict['customfield_10105'].append(watcher_dict)
            #print("========================")
            #print(watcher.strip())
    else :
        del dissue_dict['customfield_10105']
        print(keyword, " = None... Skip")

def setDuedate(keyword, value) :
    if(value is not None):
        #duedate = datetime.strptime(value, '%Y-%m-%d')
        print(keyword, " = ", value)
        dissue_dict[keyword] = str(value)
    else :
        del dissue_dict[keyword]
        print(keyword, " = None... Skip")

def setLabels(keyword, value) :
    if(value is not None):
        label_list = value.split(',')
        print(label_list)
        for label in label_list :
            dissue_dict[keyword].append(label)
    else :
        del dissue_dict[keyword]
        print(keyword, " = None... Skip")

def setComment(keyword, value) :
    if(value is not None):
        print(keyword, " = ", value)
        #dissue_dict['comment']['comments']['0']['body'] = value
        #'comment' : { 'comments' : [ { 'body' : ''}, ] }, #comment
    else :
        print(keyword, " = None... Skip")

def setAttachment(keyword, value) :
    if(value is not None):
        print(keyword, " = ", value)
    else :
        print(keyword, " = None... Skip")

def setCommonNotice(keyword, value) :
    if(value is not None):
        print(keyword, " = ", value)
        desc = dissue_dict['description']
        desc = str(desc) + '\\n\\n' + str(value)
        dissue_dict['description'] = desc
        #print("========================")
        #print(dissue_dict['description'])
        #print("========================")
    else :
        print(keyword, " = None... Skip")


def makeDevJiraJSON(key, value) :
    if (key == 'project') :
        setProject(key, value)
    elif (key == 'components'):
        setComponents(key, value)
    elif (key == 'issuetype'):
        print(key + ' = ' + value)
        setIssueType(key, value)
    elif (key == 'parent'):
        setParent(key, value)
    elif (key == 'Epic Name'):
        setEpicName(key, value)
    elif (key == 'summary'):
        setSummarynDescription(key, value)
    elif (key == 'description'):
        setSummarynDescription(key, value)
    elif (key == 'assignee'):
        setAssigneenReporter(key, value)
    elif (key == 'reporter'):
        setAssigneenReporter(key, value)
    elif (key == 'watcher'):
        setWatchers(key, value)
    elif (key == 'duedate'):
        setDuedate(key, value)
    elif (key == 'labels'):
        setLabels(key, value)
    elif (key == 'comment'):
        setComment(key, value)
    elif (key == 'attachment'):
        print("Set attachment")
    elif (key == 'Common Notice'):
        setCommonNotice(key, value)
    else :
        print("[Skip] Column - Unregistered key or field = ", key)


form_class = uic.loadUiType("./QtUI/MainDialog.ui")[0];

#class MyWindow(QMainWindow):
class MyWindow(QMainWindow, form_class) :
    def __init__(self):
        super().__init__();
        self.setupUi(self);

        self.setWindowTitle("DevTracker Jira Creator from Excel");
        self.Path.setText("Please select Excel file..........")
        self.Path.setReadOnly(True)

        self.ProgressBar.setMinimum(0)
        self.ProgressBar.setMaximum(100)
        self.ProgressBar.setValue(0)

        self.FileSelectionBtn.setEnabled(False)
        self.CreateBtn.setEnabled(False)

        self.UserID.setFocus()

        self.LoginBtn.clicked.connect(self.userLogin)
        self.CreateBtn.clicked.connect(self.createJiraIssue)
        self.ExitBtn.clicked.connect(appexit)
        self.FileSelectionBtn.clicked.connect(self.openFileNameDialog)
        #self.SaveInfo.clicked.connect(self.SaveInfoClicked)


    def userLogin(self) :
        global dev_jira
        global q_jira
        print("userLogin")
        userID = self.UserID.text()
        userPasswd = self.Passwd.text()
        try :
            dev_jira = JIRA(DevTracker, basic_auth = (userID, userPasswd))
            q_jira = JIRA(QTracker, basic_auth = (userID, userPasswd))
        except JIRAError as e:
            if e.status_code == 401 :
                print("[Error] Login Fail.. Please Check ID/Passwd and Try again!")
            print(e)
        else :
            Islogin = True
            self.FileSelectionBtn.setEnabled(True)
            self.CreateBtn.setEnabled(False)
            print("userLogin=", Islogin)
        finally :
            print("Login Try/Exception routine is passed!")
            pass


    def openFileNameDialog(self):
        global wsheet
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self,"QFileDialog.getOpenFileName()", "","Exel Files (*.xlsm);;Exel Files (*.xls)", options=options)
        if fileName:
            self.Path.setText(fileName)
            excel_file = xlsrd.load_workbook(fileName)
            wsheet = excel_file['Dev Tracker']
            Isexcelloaded = True
            self.FileSelectionBtn.setEnabled(True)
            self.CreateBtn.setEnabled(True)
            self.ProgressBar.setValue(0)

            #print("==========================================")
            #print(jira_keylist)
            #print("==========================================")
        else :
            Isexcelloaded = False
            self.CreateBtn.setEnabled(False)

    '''
    def openFileNamesDialog(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        files, _ = QFileDialog.getOpenFileNames(self,"QFileDialog.getOpenFileNames()", "","All Files (*);;Python Files (*.xlsm)", options=options)
        if files:
            print(files)

    def saveFileDialog(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getSaveFileName(self,"QFileDialog.getSaveFileName()","","All Files (*);;Text Files (*.xlsm)", options=options)
        if fileName:
            print(fileName)

    def SaveInfoClicked(self) :
    	print("SaveInfoClicked")
    '''

    def createJiraIssue(self) :
        print("createJiraIssue")
        global dissue_dict
        global dev_jira
        global q_jira

        self.ProgressBar.setValue(0)
        i = 1; j = 1
        rows = wsheet.rows
        jira_keylist = makeKeyList(wsheet)

        row_count = wsheet.max_row
        col_count = wsheet.max_column
        print('wsheet.max_row =', row_count)
        print('wsheet.max_col =', col_count)

        if (os.path.isfile("logfile.txt")) :
            os.remove("logfile.txt")

        log = open('logfile.txt', 'wt')

        '''
        update_dict = {
            'customfield_10105' :[ {"name":"" },], #watchers
        }
        '''
        for row in rows :
            if(i == 1) : i+=1; j = 1; continue

            if(wsheet.cell(row = i, column = 2).value) : # if project is not null - create issue
                dissue_dict = copy.deepcopy(dissue_init_dict)
                msg = "\n######## Row = %d Creating issues. #######\n" % i
                print(msg)
                log.write(msg)

                for key in jira_keylist :
                    if(key == 'key') :
                        j = 2; continue
                    if(key == 'Common Notice') :
                        val = wsheet.cell(row = 2, column = j).value
                    else :
                        val = wsheet.cell(row = i, column = j).value
                    #print("=====================")
                    #print(key, val)
                    #print("=====================")
                    makeDevJiraJSON(key, val)
                    j += 1
                try :
                    print("========================================================")
                    print(dissue_dict)
                    log.write(str(dissue_dict))
                    new_dissue = dev_jira.create_issue(fields= dissue_dict)
                    #dev_jira.revmove_watcher(new_dissue, 'sungbin.na')
                    new_dissue.update(fields=update_dict)
                    #createdkey = new_dissue.raw['key']
                    #print("Created Key = ", createdkey)
                    #wsheet.cell(row = i, column = 1).value = createdkey
                    print("========================================================")
                except Exception as e:
                    log.write(str(e))
                    print(e)

            else :
                msg = "\nRow = %d��° Issue ���� Skip �մϴ�.\n" % i
                print(msg)
                log.write(msg)
                pass

            progressing = int(i*100/row_count)
            self.ProgressBar.setValue(progressing)
            self.ProgressBar.update()
            i += 1

        log.close()


def appexit() :
    print("appexit")
    sys.exit()


if __name__ == "__main__" :
	app = QApplication(sys.argv);
	myWindow = MyWindow();
	myWindow.show();
	app.exec_();
