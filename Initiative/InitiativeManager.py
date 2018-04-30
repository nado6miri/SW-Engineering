import sys
import os
import copy
import time

#for Jira control
#from jira import JIRA
#from jira.exceptions import JIRAError

#pip install XlsxWriter

#xlsxwriter
#https://xlsxwriter.readthedocs.io/index.html

#openpyxl
#http://www.hanul93.com/openpyxl-basic/
#https://openpyxl.readthedocs.io/en/stable/defined_names.html



#http://hlm.lge.com/issue/rest/api/2/issue/GSWDIM-22476/
#http://hlm.lge.com/issue/rest/api/2/issue/TVPLAT-3963/
#http://hlm.lge.com/issue/rest/api/2/issue/TVPLAT-3963/editmeta
#http://hlm.lge.com/qi/rest/api/2/issue/QEVENTSEVT-7232/ - Q

import openpyxl as xlsrd
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color

from datetime import datetime
import copy

myfont = Font(name='맑은 고딕', size = 10, bold=True)
myalignment = Alignment(horizontal='center', vertical='center')
myfill = PatternFill(patternType='solid', fgColor=Color('FFC000'))
myborder = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))



DevTracker = 'http://hlm.lge.com/issue'
QTracker = 'http://hlm.lge.com/qi'

startSP = 'TVSP16_1'
endSP = 'TVSP17_2'
updateSP = 'TVSP18_1'

default_Sprint_Info = {
    'TVSP16_1' : '',  'TVSP16_2' : '',  'TVSP17_1' : '',  'TVSP17_2' : '',
    'TVSP18_1' : '',  'TVSP18_2' : '',  'TVSP19_1' : '',  'TVSP19_2' : '',
    'TVSP20_1' : '',  'TVSP20_2' : '',  'TVSP21_1' : '',  'TVSP21_2' : '',
    'TVSP22_1' : '',  'TVSP22_2' : '',  'TVSP23_1' : '',  'TVSP23_2' : '',
    'TVSP24_1' : '',  'TVSP24_2' : '',  'TVSP25_1' : '',  'TVSP25_2' : '',
    'TVSP26_1' : '',  'TVSP26_2' : '',  'TVSP27_1' : '',  'TVSP27_2' : '',
    'TVSP28_1' : '',  'TVSP28_2' : '',  'TVSP29_1' : '',  'TVSP29_2' : '',
    'TVSP30_1' : '',  'TVSP30_2' : '',  'TVSP31_1' : '',  'TVSP31_2' : '',
    }

default_epic_info = {
        'Epic Key' : '',
        'Release_SP' : '',
        'Summary' : "",
        'Assignee' : '',
        'duedate' : '',
        'Status' : '',
        'CreatedDate' : '',
        'TVSP' : { },
        'StroyCnt': '',
        'StoryResolutionRate' : '',
    }

default_initiative_info = {
    'Initiative Key' : '',
    'Summary' : '',
    'Assignee' : '',
    'Status' : '',
    'Release_SP' : '',
    'CreatedDate' : '',
    '관리대상' : '',
    'Risk 관리 대상' : '',
    'Initiative Order' : '',
    'EPIC' : [],
    'DEMO' : [],
    'CCC' : [],
    'TestCase' : [],
    'Dev_Verification' : [],
    'TVSP' : {},
    'Status Color' : '',
    'SE_Delivery' : '',
    'SE_Quality' : '',
    'ScopeOfChange' : '',
    'EpicCnt' : '',
    'EpicResolutionRate' : '',
    'StoryResolutionRate' : '',
        }


finalInfo = []

#####################################################################################################################
# JIRA Control
maxResultCnt = 3000

#===========================================================================
# Get filtered issue with Filter ID in Dev Tracker
# filter 지정을 통한 Initiative 검색
#===========================================================================
def getFilterIDResult(jiraHandle, filterId, getFieldList=()) :
    setFilterID = 'filter = ' + str(filterId)
    print('strFilterID =', setFilterID)
    #resultIssue = jiraHandle.search_issues(setFilterID, startAt = 0, maxResults = 1000, fields = setfield, expand=None)
    resultIssue = jiraHandle.search_issues(setFilterID, startAt = 0, maxResults = maxResultCnt, fields = getFieldList, expand=None)
    print("[Tracker] Get JIRA Issue with Specific Filter ID: " + setFilterID)
    return resultIssue


#===========================================================================
# Get filtered issue with Filter ID in Dev Tracker
# Query 지정을 통한 Initiative 검색
#===========================================================================
def getFilterQueryResult(jiraHandle, filterQuery, getFieldList=None) :
    # Get Filtered issue with JQL Querfy String in Q/Dev Tracker
    #setFilter = 'Filter in (M3.LK61.EU.QA1, M3.LK61.EU.QA2, M3.LK61.EU.QA3, M3.LK61.EU.QA4)'
    resultIssue = jiraHandle.search_issues(filterQuery, startAt = 0, maxResults = maxResultCnt, fields = getFieldList, expand=None)
    print("[Tracker] Get JIRA Issue with Specific Filter String: " + filterQuery)
    return resultIssue


#===========================================================================
# Get Key of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getKey(jiraIssue) :
    value = jiraIssue.raw['key']
    if(value != None) :
        #print("key = ", value)
        return value

    #print("key = Null")
    return None


#===========================================================================
# Get Summary of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getSummary(jiraIssue) :
    value = jiraIssue.raw['fields']['summary']
    if(value != None) :
        #print("summary = ", value)
        return value

    #print("summary = Null")
    return None


#===========================================================================
# Get Status of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getStatus(jiraIssue) :
    value = jiraIssue.raw['fields']['status']['name']
    #print("status = ", value)
    return value


#===========================================================================
# Get issuetype of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getIssuetype(jiraIssue) :
    value = jiraIssue.raw['fields']['issuetype']['name']
    #print("issuetype = ", value)
    return value


#===========================================================================
# Get resolution of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getResolution(jiraIssue) :
    value = jiraIssue.raw['fields']['resolution']
    #print("resolution = ", value)
    return value


#===========================================================================
# Get components of jira
# [param] jiraIssue : json object of jira
# [return] components[]
#===========================================================================
def getComponents(jiraIssue) :
    value = jiraIssue.raw['fields']['components']

    if(value != None) :
        #print("components = ", json.dumps(value))
        return value

    #print("components = Null")
    return None


#===========================================================================
# Get Release Sprint of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getReleaseSprint(jiraIssue) :
    value = jiraIssue.raw['fields']['customfield_15926']
    if(value != None) :
        #print("Release Sprint = ", value)
        return value

    #print("Release Sprint = Null")
    return None


#===========================================================================
# convert duedate to Sprint
# [param] duedate : duedate
# [return] Sprint str
#===========================================================================
def conversionDuedateToSprint(duedate) :
    if(duedate >= '' and duedate <= '') :
        return 'TVSP11'
    if(duedate >= '' and duedate <= '') :
        return 'TVSP12'
    if(duedate >= '' and duedate <= '') :
        return 'TVSP13'
    if(duedate >= '' and duedate <= '') :
        return 'TVSP14'
    if(duedate >= '' and duedate <= '') :
        return 'TVSP15'
    if(duedate >= '' and duedate <= '') :
        return 'TVSP16'
    elif(duedate >= '' and duedate <= '') :
        return 'TVSP17'
    elif(duedate >= '' and duedate <= '') :
        return 'TVSP18'
    elif(duedate >= '' and duedate <= '') :
        return 'TVSP19'
    elif(duedate >= '' and duedate <= '') :
        return 'TVSP20'
    elif(duedate >= '' and duedate <= '') :
        return 'TVSP21'
    elif(duedate >= '' and duedate <= '') :
        return 'TVSP22'
    elif(duedate >= '' and duedate <= '') :
        return 'TVSP23'
    elif(duedate >= '' and duedate <= '') :
        return 'TVSP24'
    elif(duedate >= '' and duedate <= '') :
        return 'TVSP25'
    elif(duedate >= '' and duedate <= '') :
        return 'TVSP26'
    elif(duedate >= '' and duedate <= '') :
        return 'TVSP27'
    elif(duedate >= '' and duedate <= '') :
        return 'TVSP28'
    elif(duedate >= '' and duedate <= '') :
        return 'TVSP29'
    elif(duedate >= '' and duedate <= '') :
        return 'TVSP30'
    elif(duedate >= '' and duedate <= '') :
        return 'TVSP31'

    return 'TVSP_UNDEF'


#===========================================================================
# Get Status Summary of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getStatusSummary(jiraIssue) :
    value = jiraIssue.raw['fields']['customfield_15710']
    if(value != None) :
        #print("Status Summary = ", value)
        return value

    #print("Status Summary = Null")
    return None


#===========================================================================
# Get Status Color of jira
# [param] jiraIssue : json object of jira
# [return] str (RGB)
#===========================================================================
def getStatusColor(jiraIssue) :
    value = jiraIssue.raw['fields']['customfield_15711']
    if(value != None) :
        #print("Status Color = ", json.dumps(value['value']))
        return value['value']

    #print("Status Color = Null")
    return None


#===========================================================================
# Get SE_Delivery of jira
# [param] jiraIssue : json object of jira
# [return] str (RGB)
#===========================================================================
def getSE_Dilivery(jiraIssue) :
    value = jiraIssue.raw['fields']['customfield_16988']['value']
    if(value != None) :
        #print("SE_Delivery = ", json.dumps(value))
        return value

    #print("SE_Delivery = Null")
    return None


#===========================================================================
# Get SE_Quality of jira
# [param] jiraIssue : json object of jira
# [return] str (RGB)
#===========================================================================
def getSE_Quality(jiraIssue) :
    value = jiraIssue.raw['fields']['customfield_16987']
    if(value != None) :
        value = value['value']
        #print("SE_Quality = ", json.dumps(value))
        return value

    #print("SE_Quality = Null")
    return None

#===========================================================================
# Get D_Comment of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getD_Comment(jiraIssue) :
    value = jiraIssue.raw['fields']['customfield_16984']
    #print("D_Comment = ", value)
    return value


#===========================================================================
# Get Q_Comment of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getQ_Comment(jiraIssue) :
    value = jiraIssue.raw['fields']['customfield_16983']
    #print("Q_Comment = ", value)
    return value


#===========================================================================
# Get STE Member List of jira
# [param] jiraIssue : json object of jira
# [return] QE[]
#===========================================================================
def getSTEList(jiraIssue) :
    value = jiraIssue.raw['fields']['customfield_15228']
    if(value != None) :
        #print("STE_List[] = ", json.dumps(value))
        return value

    #print("STE_List[] = Null")
    return None


#===========================================================================
# Get Initiative Order of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getInitiativeOrder(jiraIssue) :
    value = jiraIssue.raw['fields']['customfield_16986']
    #print("Initiative Order = ", value)
    return value


#===========================================================================
# Get Initiative Score of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getInitiativeScore(jiraIssue) :
    value = jiraIssue.raw['fields']['customfield_16985']
    #print("Initiative Score = ", value)
    return value

#===========================================================================
# Get Created Date of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getCreatedDate(jiraIssue) :
    value = jiraIssue.raw['fields']['created']
    #print("Created Date = ", value)
    return value


#===========================================================================
# Get Updated Date of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getUpdatedDate(jiraIssue) :
    value = jiraIssue.raw['fields']['updated']
    #print("Updated Date = ", value)
    return value


#===========================================================================
# Get Due Date of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getDueDate(jiraIssue) :
    value = jiraIssue.raw['fields']['duedate']
    #print("Due Date = ", value)
    return value

#===========================================================================
# Get Resolution Date List of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getResolutionDate(jiraIssue) :
    value = jiraIssue.raw['fields']['resolutiondate']
    #print("Resolutiondate Date = ", value)
    return value


#===========================================================================
# Get Created Date List of jira
# [param] jiraIssue : json object of jira
# [return] labels []
#===========================================================================
def getLabels(jiraIssue) :
    value = jiraIssue.raw['fields']['labels']
    #print("labels = ", value)
    return value

#===========================================================================
# Get Description List of jira
# [param] jiraIssue : json object of jira
# [return] labels [ 'a', 'b', .... ]
#===========================================================================
def getDescription(jiraIssue) :
    value = jiraIssue.raw['fields']['description']
    #print("description = ", value)
    return value

#===========================================================================
# Get fixVersions of jira
# [param] jiraIssue : json object of jira
# [return] fixVersions [ { 'name' : '' } ]
#===========================================================================
def getFixVersions(jiraIssue) :
    value = jiraIssue.raw['fields']['fixVersions']
    #print("fixVersions = ", value)
    return value


#===========================================================================
# Get Scope of Change of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getScopeOfChange(jiraIssue) :
    value = jiraIssue.raw['fields']['customfield_15104']
    if(value != None) :
        #print("Scope of Change = ", value['value'])
        return value['value']

    #print("Scope of Change = Null")
    return None


#===========================================================================
# Get Issue Links of jira
# [param] jiraIssue : json object of jira
# [return] issuelinks[ {}, .... ]
#===========================================================================
def getIssueLinks(jiraIssue) :
    value = jiraIssue.raw['fields']['issuelinks']
    #print("Issue Links = ", value)
    return value

#===========================================================================
# Get Reporter of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getReporter(jiraIssue) :
    value = jiraIssue.raw['fields']['reporter']['name']
    #print("reporter = ", value)
    return value


#===========================================================================
# Get Assignee of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getAssignee(jiraIssue) :
    value = jiraIssue.raw['fields']['assignee']['name']
    #print("assignee = ", value)
    return value


#===========================================================================
# Get Reporter of jira
# [param] jiraIssue : json object of jira
# [return] Watchers [ ] <== [ { 'name' : ''}, { 'emailAddress' : '' }, .... ]
#===========================================================================
def getWatchers(jiraIssue) :
    watchers = jiraIssue.raw['fields']['customfield_10105']
    results = []
    for watcher in watchers :
        results.append(watcher)

    #print("Watcher List = ", results)
    return results

#===========================================================================
# Get Epics / Milestone block from Description of jira
# [param] description : description string
# [param] fieldtitle : Title of block like '*Milestone*'
# [return] str or None
#===========================================================================
strEpic = ["*개발 Epic 산정*", "*Epics*", "*EPICs*" ]
strMilestone = [ "*Milestone", "*Expected Deliveries*" ]

def getEpicsMilestoneFromDesc(description, fieldtitle) :
    startpos = description.find(fieldtitle)
    if(startpos > 1) :
        endpos = description.find('*', startpos + len(fieldtitle))

    if(startpos >= 1 and endpos > startpos):
        result = description[startpos:endpos]
        #print(result)
        return result

    #print("getEpicsFromDesc = Null")
    return None

#####################################################################################################################
# Exel Control

#===========================================================================
# Get row count of excel (with Data)
# [param] Sheetname : Excel Sheet handle
# [param] rowpos : start row position of Data ( 1 ~ XXXX )
# [param] colpos : reference col index to detect exact row count (This column should be filled with data)
# [return] row count
#===========================================================================
def getRowCount(Sheetname, rowpos, colpos) :
    for i in range(rowpos, 50000) :
        val = Sheetname.cell(row = i, column = colpos).value
        if(val == None) :
            break;

    print("Row Count of ", Sheetname, " = ", i-rowpos)
    return (i-rowpos)


#===========================================================================
# Get Column count of excel (with Data)
# [param] Sheetname : Excel Sheet handle
# [param] rowpos : start row position of Title / Header
# [param] colpos : start col position of title
# [return] column count
#===========================================================================
def getColumnCount(Sheetname, rowpos, colpos) :
    for i in range(colpos,50000) :
        val = Sheetname.cell(row = rowpos, column = i).value
        if(val == None) :
            break;

    print("Column Count of ", Sheetname, " = ", i-colpos)
    return (i-colpos)



#===========================================================================
# Get Column Index of excel (with title)
# [param] Sheetname : Excel Sheet handle
# [param] rowpos : start row position of Title / Header
# [return] column Index
#===========================================================================
def getColumnIndex(Sheetname, row, title) :
    index = 1
    for col in range(1, Sheetname.max_column) :
        if(title == str(Sheetname.cell(row = row, column = col).value)) :
            break;
        index += 1

    if (Sheetname.max_column <= index) :
        index = None
        #print("title = ", title, " : Can't find title in exel")
    else :
        #print("title = ", title, ", Index = ", index)
        pass

    return (index)



#===========================================================================
# Get row Index of initiative to find
# [param] Sheetname : Excel Sheet handle
# [param] InitiativeKey : InitiativeKey to find
# [return] row Index or 0 (Not found)
#===========================================================================
def getInitiativeRowIndex(Sheetname, InitiativeKey) :
    isFound = False
    for row_index in range(1, MAX_RowCount+1) :
        type = str(Sheetname.cell(row = row_index, column = CI_IssueType).value).strip()
        if(type == "Initiative") :
            keyvalue = str(Sheetname.cell(row = row_index, column = CI_InitKey).value).strip()
            if(keyvalue == InitiativeKey) :
                isFound = True
                #print("Found Initiative Key of ", Sheetname, " : Key = ", InitiativeKey, ", Index = ", row_index)
                break;

    if(isFound == False) :
        #print("Not Found Initiative Key of ", Sheetname, " : Key = ", InitiativeKey, ", Index = ", row_index)
        row_index = 0

    return row_index




#===========================================================================
# Get row Index of Epic to find
# [param] Sheetname : Excel Sheet handle
# [param] InitiativeKey : Epic to find
# [return] row Index or 0 (Not found)
#===========================================================================
def getEpicRowIndex(Sheetname, EpicKey) :
    isFound = False
    for row_index in range(1, MAX_RowCount+1) :
        type = str(Sheetname.cell(row = row_index, column = CI_IssueType).value).strip()
        if(type == "EPIC") :
            keyvalue = str(Sheetname.cell(row = row_index, column = CI_EpicKey).value).strip()
            if(keyvalue == EpicKey) :
                isFound = True
                #print("Found Epic Key of ", Sheetname, " : Key = ", EpicKey, ", Index = ", row_index)
                break;

    if(isFound == False) :
        #print("Not Found Epic Key of ", Sheetname, " : Key = ", EpicKey, ", Index = ", row_index)
        row_index = 0

    return row_index


#===========================================================================
# Get title list of header / title
# [param] Sheetname : Excel Sheet handle
# [param] row : start row position of Title / Header
# [param] col : start col position of title
# [return] title[]
#===========================================================================
def getTitleListfromXls(Sheetname, row, col) :
    title = []
    for i in range(col, MAX_ColCount+1) :
        title.append(str(Sheetname.cell(row = row, column = i).value).strip())

    #print("Title List of ", Sheetname, " = ", title)
    return title



#===========================================================================
# Get All Initiative Key List from Excel
# [param] Sheetname : Excel Sheet handle
# [param] rowpos : start row position of Data ( 1 ~ XXXX )
# [return] Initiative Key List[ 'KEY1', 'KEY2', ... ]
#===========================================================================
def getInitiativeKeylistFromXls(Sheetname, row) :
    initative_key = []
    for row_index in range(row, MAX_RowCount+1) :
        type = str(Sheetname.cell(row = row_index, column = CI_IssueType).value).strip()
        if(type == "Initiative") :
            initative_key.append(str(Sheetname.cell(row = row_index, column = CI_InitKey).value).strip())

    #print("Initiative Key List of ", Sheetname, " = ", initative_key)
    return initative_key


#===========================================================================
# Get All Epic Key List from Excel
# [param] Sheetname : Excel Sheet handle
# [param] rowpos : start row position of Data ( 1 ~ XXXX )
# [return] Epic Key List[ 'KEY1', 'KEY2', ... ]
#===========================================================================
def getEpicKeyListfromXls(Sheetname, row) :
    epic_key = []
    for row_index in range(row, MAX_RowCount+1) :
        type = str(Sheetname.cell(row = row_index, column = CI_IssueType).value).strip()
        if(type == "EPIC") :
            epic_key.append(str(Sheetname.cell(row = row_index, column = CI_EpicKey).value).strip())

    epic_key = RemoveDuplicateInList(epic_key)
    #print("Epic Key List of ", Sheetname, " = ", epic_key)
    return epic_key



#===========================================================================
# Get Sprint History (dict) from Excel
# [param] Sheetname : Excel Sheet handle
# [param] KeyID : Initiative or Epic Key to get Sprint History from excel
# [return] Sprint_Info{ 'SP16_1' : '', 'SP16_2' : '', .... }
#===========================================================================
def getSprintHistoryfromXls(Sheetname, KeyID, IssueType) :
    Sprint_Info = { }
    Sprint_Info = copy.deepcopy(default_Sprint_Info)

    if(IssueType == "Initiative") :
        rowIndex = getInitiativeRowIndex(Sheetname, KeyID)
    elif (IssueType == "EPIC") :
        rowIndex = getEpicRowIndex(Sheetname, KeyID)
    else :
        pass

    if(rowIndex == 0) :
        pass
    else :
        for col_index in range(CI_StartPos, CI_EndPos+1) :
            tvspstr = str(Sheetname.cell(row = 2, column = col_index).value).strip()
            Sprint_Info[tvspstr] = str(Sheetname.cell(row = rowIndex, column = col_index).value).strip()

        if(IssueType == "Initiative") :
            #print("===============Update Initiative Sprint_Info===================== Row Index =", rowIndex)
            pass
        elif (IssueType == "EPIC") :
            #print("===============Update Epic Sprint_Info===================== Row Index =", rowIndex)
            pass
    #print (Sprint_Info)
    return Sprint_Info



#===========================================================================
# Get Epic Information from Excel
# [param] Sheetname : Excel Sheet handle
# [param] EpicKey : Epic Key to get detail information from excel
# [return] epic_info { 'key' : '', 'summary' : '', Sprint_Info{ 'SP16_1' : '', 'SP16_2' : '', .... } }
#===========================================================================
def getEpicInfofromXls(Sheetname, EpicKey) :
    epic_info = { }
    for row_index in range(1, MAX_RowCount+1) :
        epic_info = copy.deepcopy(default_epic_info)
        getEpicKey = str(Sheetname.cell(row = row_index, column = CI_EpicKey).value).strip()
        epic_info["Release_SP"] = str(Sheetname.cell(row = row_index, column = CI_ReleaseSP).value).strip()

        if(getEpicKey == EpicKey) :
            epic_info['Epic Key'] = getEpicKey
            spInfo = getSprintHistoryfromXls(cur_sheet, getEpicKey, "EPIC")
            epic_info['TVSP'] = spInfo

    return epic_info



#===========================================================================
# Get All Initiative - Epic Lists from Excel
# [param] Sheetname : Excel Sheet handle
# [return] epic_key [ { 'key' : 'Initative Key',  'epiclist' : [ 'Epic Key1', 'Epic Key2', ... ]}, ....  }
#===========================================================================
def getInitiativeAllEpicsListfromXls(Sheetname) :
    epic_key = []
    keylist = getInitiativeKeylistFromXls(Sheetname, 3)

    for keyID in keylist :
        tmp = { 'key' : '', 'epiclist' : []}
        tmp['key'] = keyID
        for row_index in range(1, MAX_RowCount+1) :
            type = str(Sheetname.cell(row = row_index, column = CI_IssueType).value).strip()
            epicparent = str(Sheetname.cell(row = row_index, column = CI_InitKey).value).strip()
            if(type == 'EPIC' and epicparent == keyID) :
                tmp['epiclist'].append(str(Sheetname.cell(row = row_index, column = CI_EpicKey).value).strip())
        epic_key.append(tmp)

    #print("*********** All Epic key List from Xls **********************")
    #print(epic_key)
    return epic_key


#===========================================================================
# Get a specific Initiative - Epic Lists from Excel
# [param] Sheetname : Excel Sheet handle
# [param] InitiativeKey : Initiative Key to get Epic Lists
# [return] EpicList[ 'Epic Key1', 'Epic Key2', ... ]
#===========================================================================
def getInitiativeEpicListsfromXls(Sheetname, InitiativeKey) :
    epic_key = []
    for row_index in range(1, MAX_RowCount+1) :
        type = str(Sheetname.cell(row = row_index, column = CI_IssueType).value).strip()
        getInitkey = str(Sheetname.cell(row = row_index, column = CI_InitKey).value).strip()
        if(type == "EPIC" and InitiativeKey == getInitkey) :
            epic_key.append(str(Sheetname.cell(row = row_index, column = CI_EpicKey).value).strip())

    #print("Initiative Key = ", InitiativeKey, " Epic Key List of ", Sheetname, " = ", epic_key)
    return epic_key



#===========================================================================
# Set value to Cell
# [param] Sheetname : Excel Sheet handle
# [param] row : row
# [param] col : column
# [return] None
#===========================================================================
def setXlsCell(Sheetname, row, col, value) :
    #Sheetname.cell(row=row, column=col).font = myfont
    #Sheetname.cell(row=row, column=col).alignment = myalignment
    #Sheetname.cell(row=row, column=col).fill = myfill
    #Sheetname.cell(row=row, column=col).border = myborder

    Sheetname.cell(row=row, column=col).value = value


#===========================================================================
# Get value from Cell
# [param] Sheetname : Excel Sheet handle
# [param] row : row
# [param] col : column
# [return] value
#===========================================================================
def getXlsCell(Sheetname, row, col) :
    value = str(Sheetname.cell(row, col).value).strip()
    return value



#===========================================================================
# Get All Initiative Key List from Jira
# [param] rowpos : start row position of Data ( 1 ~ XXXX )
# [return] Initiative Key List[ 'KEY1', 'KEY2', ... ]
#===========================================================================
def getInitiativeKeylistFromJira(filterResult) :
    initative_key = []
    for issue in filterResult :
        initiative_key.append(getKey(issue))

    #print("Initiative Key List from Jira = ", initative_key)
    return initative_key



#===========================================================================
# Get All Initiative - Epic Lists from Jira
# [param] filterResult : Jira Result from Filtered JIRA Query
# [return] list[ { 'key' : 'Initative Key',  'epiclist' : [ 'Epic Key1', 'Epic Key2', ... ]}, ....  ]
#===========================================================================
def getInitiativeAllEpicsListfromJira(filterResult) :
    epic_key = []

    for dissue in filterResult :
        # Get issue with All Fields in Dev Tracker
        tmp = { 'key' : '', 'epiclist' : []}
        tmp['key'] = dissue.raw['key']
        bfound = False
        for issuelink in dissue.raw['fields']['issuelinks'] :
            if 'outwardIssue' in issuelink :
                if(issuelink['outwardIssue']['fields']['issuetype']['name'] == 'Epic') :
                    #print ("Key = ", dissue.raw['key'], " Status = ", issuelink['outwardIssue']['fields']['status']['name'], " Linked Issue = ", issuelink['outwardIssue']['key'])
                    tmp['epiclist'].append(issuelink['outwardIssue']['key'])
                    bfound = True
            if 'inwardIssue' in issuelink :
                if(issuelink['inwardIssue']['fields']['issuetype']['name'] == 'Epic') :
                    #print ("Key = ", dissue.raw['key'], " Status = ", issuelink['inwardIssue']['fields']['status']['name'], " Linked Issue = ", issuelink['inwardIssue']['key'])
                    tmp['epiclist'].append(issuelink['inwardIssue']['key'])
                    bfound = True

        if(bfound == True) :
            epic_key.append(tmp)

    #print("*********** Initiative - All Epic key List **********************")
    #print(epic_key)
    return epic_key



#===========================================================================
# Get All Epic Lists from Jira
# [param] filterResult : Jira Result from Filtered JIRA Query (Initiative Filter)
# [return] epic_key[ 'Epic Key1', 'Epic Key2', ... ]
#===========================================================================
def getEpicKeyListfromJira(filterResult, rawData) :
    epic_key = []

    if(rawData == "Initiative_Filter") : # make a epic list from Issuelinks
        for dissue in filterResult :
            # Get issue with All Fields in Dev Tracker
            for issuelink in dissue.raw['fields']['issuelinks'] :
                if 'outwardIssue' in issuelink :
                    if(issuelink['outwardIssue']['fields']['issuetype']['name'] == 'Epic') :
                        #print ("Key = ", dissue.raw['key'], " Status = ", issuelink['outwardIssue']['fields']['status']['name'], " Linked Issue = ", issuelink['outwardIssue']['key'])
                        epic_key.append(issuelink['outwardIssue']['key'])
                if 'inwardIssue' in issuelink :
                    if(issuelink['inwardIssue']['fields']['issuetype']['name'] == 'Epic') :
                        #print ("Key = ", dissue.raw['key'], " Status = ", issuelink['inwardIssue']['fields']['status']['name'], " Linked Issue = ", issuelink['inwardIssue']['key'])
                        epic_key.append(issuelink['inwardIssue']['key'])
    elif (rawData == "Epic_Filter") : # make a epic list from Epic Filter Result
        # Compare Epic List ........
        for dissue in filterResult :
            epic_key.append(getKey(dissue))
    else :
        pass

    # remove duplicate item in list
    epic_key = RemoveDuplicateInList(epic_key)

    #print("*********** All Epic key List from Jira (rawData = {0})**********************".format(rawData))
    #print(epic_key)
    return epic_key


#===========================================================================
# Get a specific Initiative - Epic Lists from Jira
# [param] jiraAllEpicList : list[ { 'key' : 'Initative Key',  'epiclist' : [ 'Epic Key1', 'Epic Key2', ... ]}, ....  }
# [param] InitiativeKey : Initiative Key to get Epic Lists
# [return] EpicList[ 'Epic Key1', 'Epic Key2', ... ]
#===========================================================================
def getInitiativeEpicListsfromJira(jiraAllEpicList, InitiativeKey) :
    epic_key = []

    for item in jiraAllEpicList :
        if (InitiativeKey == item['key']) :
            epic_key.append(item['epiclist'])

    #print("*********** Initiative Key = {0} Epic key from JIRA **********************".format(InitiativeKey))
    #print(epic_key)
    return epic_key



#===========================================================================
# Get the detail Initiative Information needed for history management from excel
# [param] Sheetname : Excel Sheet handle
# [param] IntiativeKeyList : Initiative List[ 'KEY1', 'KEY2', ... ]
# [param] Init_EpicList : Initative - Epic Info list[ { 'key' : 'Initative Key',  'epiclist' : [ 'Epic Key1', 'Epic Key2', ... ]}, ....  ]
# [return] all initiative_info list []
#===========================================================================
def getInitiativeDetailInfofromXls(Sheetname, IntiativeKeyList, Init_EpicList) :
    result = []
    initiative_info = {}

    for key in IntiativeKeyList :
        initiative_info = copy.deepcopy(default_initiative_info)
        rowIndex = getInitiativeRowIndex(Sheetname, key)

        if(rowIndex > 0) :
            print("\n######## {0} row - Update Initiative Detail information from Xls".format(rowIndex))
            initiative_info["Initiative Key"] = str(Sheetname.cell(row = rowIndex, column = CI_InitKey).value).strip()
            initiative_info["Release_SP"] = str(Sheetname.cell(row = rowIndex, column = CI_ReleaseSP).value).strip()
            initiative_info["관리대상"] = str(Sheetname.cell(row = rowIndex, column = 2).value).strip()
            initiative_info["Risk 관리 대상"] = str(Sheetname.cell(row = rowIndex, column = 3).value).strip()

            #SP
            spInfo = getSprintHistoryfromXls(cur_sheet, key, "Initiative")
            initiative_info['TVSP'] = spInfo

            #EPIC
            epic_list = getInitiativeEpicListsfromJira(Init_EpicList, key)
            for epickey in epic_list :
                # epic_info = { 'Epic Key' : '', 'Summary' : '', Sprint_Info{ 'SP16_1' : '', 'SP16_2' : '', .... } }
                epicInfo = getEpicInfofromXls(cur_sheet, epickey)
                initiative_info['EPIC'].append(epicInfo)

            result.append(initiative_info)
            #print(initiative_info)

    return result



#===========================================================================
# Get the detail Initiative Information needed for history management from Jira
# [param] finalinfo : Initative Detail Info after updateing data from excel.
#   final_info = [ default_initiative_info1, default_initiative_info2, default_initiative_info3, ..., default_initiative_infoN ]
#   default_initiative_info = {
#        'Initiative Key' : '',
#        'Summary' : '',
#        'Assignee' : '',
#        'Status' : '',
#        'Release_SP' : '',
#        'CreatedDate' : '',
#        '관리대상' : '',
#        'Risk 관리 대상' : '',
#        'Initiative Order' : '',
#        'EPIC' : [],
#        'DEMO' : [],
#        'CCC' : [],
#        'TestCase' : [],
#        'Dev_Verification' : [],
#        'TVSP' : {},
#        }
#
#    default_epic_info = {
#            'Epic Key' : '',
#            'Release_SP' : '',
#            'Summary' : "",
#            'Assignee' : '',
#            'duedate' : '',
#            'Status' : '',
#            'CreatedDate' : '',
#            'TVSP' : { },
#        }
#
# [return] Final Data : all initiative_info list []
#===========================================================================
def getInitiativeDetailInfofromJira(Initiative_FilterResult, Epic_FilterResult, finalinfo) :

    for initiative in finalinfo :
        dissue = getJiraInfo(Initiative_FilterResult, initiative['Initiative Key'])
        if(dissue != None) :
            initiative['Summary'] = getSummary(dissue)
            initiative['Assignee'] = getAssignee(dissue)
            initiative['Status'] = getStatus(dissue)
            initiative['CreatedDate'] = getCreatedDate(dissue)
            initiative['Initiative Order'] = getInitiativeOrder(dissue)
            initiative['Status Color'] = getStatusColor(dissue)
            initiative['SE_Delivery'] = getSE_Dilivery(dissue)
            initiative['SE_Quality'] = getSE_Quality(dissue)
            initiative['ScopeOfChange'] = getScopeOfChange(dissue)
            initiative['TVSP'][updateSP] = getReleaseSprint(issue) #duedate 기반 SP 정보 기입 (Release Sprint 값 적용)
            initiative['EpicCnt'] = len(initiative['EPIC'])
            initiative['EpicResolutionRate'] = 0
            initiative['StoryResolutionRate'] = 0

            for epic in initiative['EPIC'] :
                epicissue = getJiraInfo(Epic_FilterResult, epic['Epic Key'])
                if(epicissue != None) :
                    epic['EPIC']['Summary'] = getSummary(epicissue)
                    epic['EPIC']['Assignee'] = getAssignee(epicissue)
                    epic['EPIC']['Status'] = getStatus(epicissue)
                    epic['EPIC']['CreatedDate'] = getCreatedDate(epicissue)
                    epic['EPIC']['TVSP'][updateSP] = conversionDuedateToSprint(epic['EPIC']['duedate'])
                    epic['EPIC']['duedate'] = getDueDate(epicissue)
                    epic['EPIC']['StroyCnt'] = 0
                    epic['EPIC']['StoryResolutionRate'] = 0
        else :
            print("Can't find Initiative or Epic Key from JIRA Filter Result.")
            pass

    return finalinfo


#===========================================================================
# Get a initiative / Epic issue from Jira
# [param] filterResult : Jira Result from Filtered JIRA Query (Initiative or Epic Filter)
# [param] InitiativeKey : Initiative or Epic Key
# [return] Initiative jira issue
#===========================================================================
def getJiraInfo(filterResult, KeyID) :
    for dissue in filterResult :
        if(KeyID == getKey(dissue)) :
            return dissue
    return None



#===========================================================================
# Get Jira Result from FilterID
# [param] jiraHandle : Dev Jira handle
# [param] filterid : filter ID
# [return] JIRA Filtered Result
#===========================================================================
def getFilteredInitiativeInfofromJira(jiraHandle, filterid) :
    #Filter ID
    Initiative_webOS45_Initial_Dev = filterid

    #setfield = ('summary, duedate, assignee, status, created, components, labels')
    #result = getFilterIDResult(jiraHandle, Initiative_webOS45_Initial_Dev, setfield)
    if(Initiative_webOS45_Initial_Dev == 42101) :
        result = getFilterIDResult(jiraHandle, Initiative_webOS45_Initial_Dev)
    else :
        setfield = ('summary, comment, assignee, duedate, created, labels')
        result = getFilterIDResult(jiraHandle, Initiative_webOS45_Initial_Dev, setfield)

    return result


#===========================================================================
# Python code to remove duplicate elements
# [param] duplicate : List with dulplicated Data
# [return] List Data
#===========================================================================
def RemoveDuplicateInList(duplicate):
    final_list = []
    for num in duplicate:
        if num not in final_list:
            final_list.append(num)
    return final_list


#===========================================================================
# Get Difference between List A and List B
# [param] listA : List Data
# [param] listB : List Data
# [return] List Data
#===========================================================================
def getDiffList(listA, listB) :
    return (list(set(listA) - set(listB)))




#===========================================================================
# Get the detail Initiative Information needed for history management from Jira
# [param] finalinfo : Initative Detail Info after updateing data from excel.
#   final_info = [ default_initiative_info1, default_initiative_info2, default_initiative_info3, ..., default_initiative_infoN ]
#   default_initiative_info = {
#        'Initiative Key' : '',
#        'Summary' : '',
#        'Assignee' : '',
#        'Status' : '',
#        'Release_SP' : '',
#        'CreatedDate' : '',
#        '관리대상' : '',
#        'Risk 관리 대상' : '',
#        'Initiative Order' : '',
#        'EPIC' : [],
#        'DEMO' : [],
#        'CCC' : [],
#        'TestCase' : [],
#        'Dev_Verification' : [],
#        'TVSP' : {},
#        }
#
#    default_epic_info = {
#            'Epic Key' : '',
#            'Release_SP' : '',
#            'Summary' : "",
#            'Assignee' : '',
#            'duedate' : '',
#            'Status' : '',
#            'CreatedDate' : '',
#            'TVSP' : { },
#        }
#
# [return] Final Data : all initiative_info list []
#===========================================================================
def updateInitiativeDetailInfoToXls(Sheetname, finalinfo) :
    row_index = 3
    col_index = 1
    index = 1
    for initiative in finalinfo :
        # write No Column
        setXlsCell(Sheetname, row_index, 1).value = index

        # write Initiative Order
        setXlsCell(Sheetname, row_index, 2).value = initiative['Initiative Order']
        # write Issue Type
        setXlsCell(Sheetname, row_index, 3).value = 'Initiative'
        # write Initiative Key
        setXlsCell(Sheetname, row_index, 4).value = initiative['Initiative Key']
        # write Epic Key
        setXlsCell(Sheetname, row_index, 5).value = initiative['Initiative Key']
        # write Summary
        setXlsCell(Sheetname, row_index, 6).value = initiative['Summary']

        # write Assignee
        setXlsCell(Sheetname, row_index, 7).value = initiative['Assignee']
        # write Status
        setXlsCell(Sheetname, row_index, 8).value = initiative['Status']
        # write CreatedDate
        setXlsCell(Sheetname, row_index, 9).value = initiative['CreatedDate']
        # write Release Sprint
        setXlsCell(Sheetname, row_index, 10).value = initiative['Release_SP']
        # write SP
        for key, value in initiative['SP'].items() :
            colpos = getColumnIndex(Sheetname, key)
            setXlsCell(Sheetname, row_index, colpos).value = value

        '''
        initiative['Status Color'] = getStatusColor(dissue)
        initiative['SE_Delivery'] = getSE_Dilivery(dissue)
        initiative['SE_Quality'] = getSE_Quality(dissue)
        initiative['ScopeOfChange'] = getScopeOfChange(dissue)
        initiative['TVSP']['updateSP'] = getReleaseSprint(issue) #duedate 기반 SP 정보 기입 (Release Sprint 값 적용)
        initiative['EpicCnt'] = len(initiative['EPIC'])
        initiative['EpicResolutionRate'] = 0
        initiative['StoryResolutionRate'] = 0
        '''

        row_index += 1
        index += 1
        for epic in initiative['EPIC'] :
            # write No Column
            setXlsCell(Sheetname, row_index, 1).value = index

            # write Initiative Order
            setXlsCell(Sheetname, row_index, 2).value = initiative['Initiative Order']
            # write Issue Type
            setXlsCell(Sheetname, row_index, 3).value = 'EPIC'
            # write Initiative Key
            setXlsCell(Sheetname, row_index, 4).value = initiative['Initiative Key']
            # write Epic Key
            setXlsCell(Sheetname, row_index, 5).value = epic['Epic Key']
            # write Summary
            setXlsCell(Sheetname, row_index, 6).value = epic['Summary']
            # write Assignee
            setXlsCell(Sheetname, row_index, 7).value = epic['Assignee']
            # write Status
            setXlsCell(Sheetname, row_index, 8).value = epic['Status']
            # write CreatedDate
            setXlsCell(Sheetname, row_index, 9).value = epic['CreatedDate']
            # write Release Sprint
            setXlsCell(Sheetname, row_index, 10).value = epic['Release_SP']
            # write SP
            for key, value in epic['SP'].items() :
                colpos = getColumnIndex(Sheetname, key)
                setXlsCell(Sheetname, row_index, colpos).value = value

            '''
            epic['EPIC']['TVSP'][updateSP] = conversionDuedateToSprint(epic['EPIC']['duedate'])
            epic['EPIC']['duedate'] = getDueDate(epicissue)
            epic['EPIC']['StroyCnt'] = 0
            epic['EPIC']['StoryResolutionRate'] = 0
            '''

            row_index += 1
            index += 1



#===========================================================================
# Clear data within region
# [param] Sheetname : Excel Sheet Name
# [param] start_row : Start Row Position of region to be cleared
# [param] start_col : Start Column Position of region to be cleared
# [return] None
#===========================================================================
def prepareNewXlsSheet(Sheetname, start_row, start_col) :
    for x in range(start_row, maxResultCnt) :
        for y in range(start_col, 50) :
            setXlsCell(Sheetname, x, y, '')



#===========================================================================
# Main Function
# [param] None
# [return] None
#===========================================================================
if __name__ == "__main__" :
    # jira Handle open
    #dev_jira = JIRA(DevTracker, basic_auth = ("sungbin.na", ""))

    # create log file
    if (os.path.isfile("Initiative_logfile.txt")) :
        os.remove("Initiative_logfile.txt")

    log = open('Initiative_logfile.txt', 'wt')

    # Create Excel workbook
    workbook = xlsrd.load_workbook('Initiative일정관리_180426_v1.xlsx')
    org_sheet = workbook["최종"]

    workbook.copy_worksheet(org_sheet)
    cur_sheet = workbook["최종 Copy"]
    #print(workbook.get_sheet_names())
    cur_sheet.title = "금일작업본"

    prepareNewXlsSheet(cur_sheet, 3, 1)

    # set max row/column count
    MAX_RowCount = getRowCount(cur_sheet, 3, 1)
    MAX_ColCount = getColumnCount(cur_sheet, 2, 1)

    # set title column index to variables
    CI_IssueType = getColumnIndex(cur_sheet, 2, "Type")
    CI_EpicKey = getColumnIndex(cur_sheet, 2, "Epic Key")
    CI_InitKey = getColumnIndex(cur_sheet, 2, "Initiative Key")
    CI_ReleaseSP = getColumnIndex(cur_sheet, 2, "Release_SP")
    CI_StartPos = getColumnIndex(cur_sheet, 2, startSP)
    CI_EndPos = getColumnIndex(cur_sheet, 2, endSP)
    CI_Summary = getColumnIndex(cur_sheet, 2, "Summary")
    CI_Assignee = getColumnIndex(cur_sheet, 2, "Assignee")
    CI_Status = getColumnIndex(cur_sheet, 2, "Status")
    CI_Created = getColumnIndex(cur_sheet, 2, "CreatedDate")
    CI_InitOrder = getColumnIndex(cur_sheet, 2, "Initiative Order")


    # Initiative ========================================================================
    # 1. JIRA에서 Initiative Filter에 맞는 Initiative Key를 Jira로 구성한다. [ 'key1', 'key2', .... ]
    Initiative_FilterResult = getFilteredInitiativeInfofromJira(dev_jira, 42101)
    jira_initiative_keylist = getInitiativeKeylistFromJira(Initiative_FilterResult)

    # 2. Excel로 부터 Initiative Key List를 구성한다. [ 'key1', 'key2', .... ]
    xls_initiative_keylist = getInitiativeKeylistFromXls(cur_sheet, 3)

    # 3. Jira상의 Initiative Key List와 엑셀상에 관리되는 Initiative Key List를 비교한다.
    print("\n################## New Initiative List (JIRA - Excel) ##################")
    newkey = getDiffList(jira_initiative_keylist, xls_initiative_keylist)
    print(newkey)
    print("\n################## Del Initiative List (Excel - JIRA) ##################")
    delkey = getDiffList(xls_initiative_keylist, jira_initiative_keylist)
    print(delkey)


    # Epic ==============================================================================
    # 4. JIRA에서 Initative IssueLinks 정보에서 Epic Key List를 구성한다. [ 'key1', 'key2', .... ]
    jira_Issuelinks_epic_keylist = getEpicKeyListfromJira(Initiative_FilterResult, "Initiative_Filter")

    # 5. JIRA에서 Epic Filter에 맞는 Epic Key List를 구성한다. [ 'key1', 'key2', .... ]
    Epic_FilterResult = getFilteredInitiativeInfofromJira(dev_jira, 42317)
    jira_epic_keylist = getEpicKeyListfromJira(Epic_FilterResult, "Epic_Filter")
    xls_epic_keylist = getEpicKeyListfromXls(cur_sheet, 3)

    # 6. JIRA에서 Epic Filter를 이용해 구성한 정보와 Initiative Issuelinks[] 정보를 이용해 만든 Epic Key List가 일치하는지를 체크한다. [ 'key1', 'key2', .... ]
    print("\n################## Compare1 Epic List (jira filter - jira link) ##################")
    new_issuelinks_epickey = getDiffList(jira_epic_keylist, jira_Issuelinks_epic_keylist)
    print(new_issuelinks_epickey)
    print("\n################## Compare2 Epic List (jira link - jira filter) ##################")
    new_filtered_epickey = getDiffList(jira_Issuelinks_epic_keylist, jira_epic_keylist)
    print(new_filtered_epickey)

    print("\n################## New Epic List (JIRA - Excel) ##################")
    newEpickey = getDiffList(jira_Issuelinks_epic_keylist, xls_epic_keylist)
    print(newEpickey)
    print("\n################## Del Epic List (Excel - JIRA) ##################")
    delEpickey = getDiffList(xls_epic_keylist, jira_Issuelinks_epic_keylist)
    print(delEpickey)

    # 7. 각 Initiative 하위에 존재하는 Epick List 구성
    # list[ { 'key' : 'Initative Key',  'epiclist' : [ 'Epic Key1', 'Epic Key2', ... ]}, ....  ]
    jira_Init_EpicLists = getInitiativeAllEpicsListfromJira(Initiative_FilterResult)


    # 8. Jira로 부터 얻은 Initiative Key를 가지고 엑셀의 정보를 먼저 Update 한다. (Jira 기준 - Initative Key List[])
    #    Update Detail Initative(Epic) Information from Excel first.
    #    (Release SP, SP History, 관리대상, 관리대상 Risk, ?Epic List?)
    tmp_Initiative = getInitiativeDetailInfofromXls(cur_sheet, jira_initiative_keylist, jira_Init_EpicLists)


    # 8. Initiative Key를 기준으로 Jira상의 최신 정보를 Update한다.
    finalInfo = getInitiativeDetailInfofromJira(Initiative_FilterResult, Epic_FilterResult, tmp_Initiative)


    # 9. Jira로 부터 얻은 Initiative Key를 기준으로 Jira상의 최신 정보를 Excel 문서에 Update 한다.
    prepareNewXlsSheet(cur_sheet, 3, 1)
    updateInitiativeDetailInfoToXls(finalInfo)

    # 10. Jira상의 최신 정보를 Excel 문서에 Update 후 별도 Excel File로 저장한다.
    workbook.save('Initiative일정관리_180426_v2.xlsx')

    '''
    #log.write(a)
    '''
