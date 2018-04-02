#pip install XlsxWriter

#xlsxwriter
#https://xlsxwriter.readthedocs.io/index.html

#openpyxl
#http://www.hanul93.com/openpyxl-basic/
#https://openpyxl.readthedocs.io/en/stable/defined_names.html

import xlsxwriter as xlswt
import openpyxl as xlsrd
from datetime import datetime


# workbook 만들기
workbook = xlswt.Workbook('hello.xlsx')

# Sheet 생성하기
ws1 = workbook.add_worksheet(name='Initiative')
ws1.write('A1', 'Hello world')
ws1.write('A2', 'Hello world')

# workbook 저장하기
workbook.close()

with xlswt.Workbook('hello_world.xlsx') as workbook:
    workbook.set_size(1200, 800)
    workbook.set_properties({
        'title':    'This is an example spreadsheet',
        'subject':  'With document properties',
        'author':   'John McNamara',
        'manager':  'Dr. Heinz Doofenshmirtz',
        'company':  'of Wolves',
        'category': 'Example spreadsheets',
        'keywords': 'Sample, Example, Properties',
        'comments': 'Created with Python and XlsxWriter'})

    worksheet = workbook.add_worksheet(name='Initiative')
    worksheet.write('A1', 'Hello world')
    worksheet.write('A2', 'Hello world')

# Global workbook name.
workbook.define_name('Sales', '=Sheet1!$G$1:$H$10')
# Local worksheet name.
workbook.define_name('Sheet2!Sales', '=Sheet2!$G$1:$G$10')
#If the sheet name contains spaces or special characters you must follow the Excel convention and enclose it in single quotes:
workbook.define_name("'New Data'!Sales", '=Sheet2!$G$1:$G$10')

# worksheet 개수만큼 반복하기
for worksheet in workbook.worksheets():
    worksheet.write('A1', 'Hello')

# get predefined worksheet
worksheet = workbook.get_worksheet_by_name('Sheet1')

# 엑셀파일 열기
excel_file = xlsrd.load_workbook('hello.xlsx')
excel_sheet = excel_file['Initiative']

for row in excel_sheet.rows :
    print(row[0].value)


excel_file.save("hello.xlsx")
excel_file.close()


#===============================================================================
# Set / Get Row and Col index.
#===============================================================================
from xlsxwriter.utility import xl_rowcol_to_cell
cell = xl_rowcol_to_cell(0, 0)   # A1
cell = xl_rowcol_to_cell(0, 1)   # B1
cell = xl_rowcol_to_cell(1, 0)   # A2
str = xl_rowcol_to_cell(0, 0, col_abs=True)                # $A1
str = xl_rowcol_to_cell(0, 0, row_abs=True)                # A$1
str = xl_rowcol_to_cell(0, 0, row_abs=True, col_abs=True)  # $A$1

(row, col) = xl_cell_to_rowcol('A1')    # (0, 0)
(row, col) = xl_cell_to_rowcol('B1')    # (0, 1)
(row, col) = xl_cell_to_rowcol('C2')    # (1, 2)
(row, col) = xl_cell_to_rowcol('$C2')   # (1, 2)
(row, col) = xl_cell_to_rowcol('C$2')   # (1, 2)
(row, col) = xl_cell_to_rowcol('$C$2')  # (1, 2)

column = xl_col_to_name(0)    # A
column = xl_col_to_name(1)    # B
column = xl_col_to_name(702)  # AAA
column = xl_col_to_name(0, False)  # A
column = xl_col_to_name(0, True)   # $A
column = xl_col_to_name(1, True)   # $B

cell_range = xl_range(0, 0, 9, 0)  # A1:A10
cell_range = xl_range(1, 2, 8, 2)  # C2:C9
cell_range = xl_range(0, 0, 3, 4)  # A1:E4

cell_range = xl_range_abs(0, 0, 9, 0)  # $A$1:$A$10
cell_range = xl_range_abs(1, 2, 8, 2)  # $C$2:$C$9
cell_range = xl_range_abs(0, 0, 3, 4)  # $A$1:$E$4

#===============================================================================
# Set Excel Style & Format with xlsxwriter
# https://xlsxwriter.readthedocs.io/format.html#format
#===============================================================================
# Add a bold format to use to highlight cells.
bold = workbook.add_format({'bold': True})

# Add a number format for cells with money.
money = workbook.add_format({'num_format': '$#,##0'})

# Add an Excel date format.
date_format = workbook.add_format({'num_format': 'mmmm d yyyy'})
cell_format = workbook.add_format({'bold': True, 'italic': True})
cell_format = workbook.add_format({'bold': True, 'font_color': 'red'})

cell_format = workbook.add_format()
cell_format.set_bold()
cell_format.set_font_color('red')


worksheet.set_row(0, 18, cell_format)
worksheet.set_row(row, options={'hidden': True})
worksheet.set_column('A:D', 20, cell_format)
worksheet1.set_selection('C3') # Set default position to C3 cell.


'''
Category    Description	   Property	    Method Name
Font        Font type	   'font_name'	set_font_name()
Font        size	       'font_size'	set_font_size()
Font        color	       'font_color'	set_font_color()
Bold	                   'bold'	    set_bold()
Italic	'italic'	set_italic()
Underline	'underline'	set_underline()
Strikeout	'font_strikeout'	set_font_strikeout()
Super/Subscript	'font_script'	set_font_script()
Number	Numeric format	'num_format'	set_num_format()
Protection	Lock cells	'locked'	set_locked()
Hide formulas	'hidden'	set_hidden()
Alignment	Horizontal align	'align'	set_align()
Vertical align	'valign'	set_align()
Rotation	'rotation'	set_rotation()
Text wrap	'text_wrap'	set_text_wrap()
Reading order	'reading_order'	set_reading_order()
Justify last	'text_justlast'	set_text_justlast()
Center across	'center_across'	set_center_across()
Indentation	'indent'	set_indent()
Shrink to fit	'shrink'	set_shrink()
Pattern	Cell pattern	'pattern'	set_pattern()
Background color	'bg_color'	set_bg_color()
Foreground color	'fg_color'	set_fg_color()
Border	Cell border	'border'	set_border()
Bottom border	'bottom'	set_bottom()
Top border	'top'	set_top()
Left border	'left'	set_left()
Right border	'right'	set_right()
Border color	'border_color'	set_border_color()
Bottom color	'bottom_color'	set_bottom_color()
Top color	'top_color'	set_top_color()
Left color	'left_color'	set_left_color()
Right color	'right_color'	set_right_color()
'''

#===============================================================================
# Write some data to Cell.
#===============================================================================
# Some data we want to write to the worksheet.
expenses = (
 ['Rent', '2013-01-13', 1000],
 ['Gas',  '2013-01-14',  100],
 ['Food', '2013-01-16',  300],
 ['Gym',  '2013-01-20',   50],
)

worksheet.write(0, 0, 'Hello', bold)    # write_string()
worksheet.write(1, 0, 'World')          # write_string()
worksheet.write(2, 0, 2)                # write_number()
worksheet.write(3, 0, 3.00001)          # write_number()
worksheet.write(4, 0, '=SIN(PI()/4)')   # write_formula()
worksheet.write(5, 0, '')               # write_blank()
worksheet.write(6, 0, None)             # write_blank()

worksheet.write_string(0, 0, 'Your text here')
worksheet.write_string('A1', 'Your text here')

worksheet.write_number(0, 0, 123456)
worksheet.write_number('A2', 2.3451)

worksheet.write_blank(0, 0, None, format)

worksheet.write_formula(0, 0, '=B3 + B4')
worksheet.write_formula(1, 0, '=SIN(PI()/4)')
worksheet.write_formula(2, 0, '=SUM(B1:B5)')
worksheet.write_formula('A4', '=IF(A3>1,"Yes", "No")')
worksheet.write_formula('A5', '=AVERAGE(1, 2, 3, 4)')
worksheet.write_formula('A6', '=DATEVALUE("1-Jan-2013")')

date_time = datetime.datetime.strptime('2013-01-23', '%Y-%m-%d')
date_format = workbook.add_format({'num_format': 'd mmmm yyyy'})
worksheet.write_datetime('A1', date_time, date_format)

worksheet.write_boolean(0, 0, True)
worksheet.write_boolean('A2', False)

worksheet.write    ('A2', 'http://www.python.org/')  # Same.
worksheet.write_url('A1', 'ftp://www.python.org/')
worksheet.write_url('A2', 'http://www.python.org/')
worksheet.write_url('A3', 'https://www.python.org/')
worksheet.write_url('A4', 'mailto:jmcnamara@cpan.org')
worksheet.write_url('A1', 'http://www.python.org', string='Python home')

# Start from the first cell below the headers.
row = 1
col = 0

for item, date_str, cost in (expenses):
    # Convert the date string into a datetime object.
    date = datetime.strptime(date_str, "%Y-%m-%d")

    worksheet.write_string  (row, col,     item              )
    worksheet.write_datetime(row, col + 1, date, date_format )
    worksheet.write_number  (row, col + 2, cost, money_format)
    row += 1


#===============================================================================
# Set Auto Filter : https://xlsxwriter.readthedocs.io/working_with_autofilters.html
# Example : https://xlsxwriter.readthedocs.io/example_autofilter.html#ex-autofilter
#===============================================================================
'x == b*'      # begins with b
'x != b*'      # doesn't begin with b
'x == *b'      # ends with b
'x != *b'      # doesn't end with b
'x == *b*'     # contains b
'x != *b*'     # doesn't contains b
'x == Blanks'
'x == NonBlanks'
'x <  2000'
'x >  2000'
'x == 2000'
'x >  2000 and x <  5000'
'x == 2000 or  x == 5000'

worksheet.autofilter('A1:D11')
worksheet.autofilter(0, 0, 10, 3)  # Same as above.

worksheet2.filter_column(0, 'Region == East')
worksheet3.filter_column('A', 'x == East or x == South')

worksheet.filter_column_list('A', ['March', 'April', 'May'])
worksheet.filter_column_list('B', [100, 110, 120, 130])


#===============================================================================
# Add Chart : https://xlsxwriter.readthedocs.io/working_with_charts.html
#===============================================================================
# Add the worksheet data to be plotted.
data = [10, 40, 50, 20, 10, 50]
worksheet.write_column('A1', data)

# Create a new chart object.
chart = workbook.add_chart({'type': 'line'})

# Add a series to the chart.
chart.add_series({'values': '=Sheet1!$A$1:$A$6'})

# Insert the chart into the worksheet.
worksheet.insert_chart('C1', chart)

#===============================================================================
# Set Freeze Panes : https://xlsxwriter.readthedocs.io/example_panes.html
#===============================================================================
workbook = xlsxwriter.Workbook('panes.xlsx')

worksheet1 = workbook.add_worksheet('Panes 1')
worksheet2 = workbook.add_worksheet('Panes 2')
worksheet3 = workbook.add_worksheet('Panes 3')
worksheet4 = workbook.add_worksheet('Panes 4')

#######################################################################
header_format = workbook.add_format({'bold': True,
                                     'align': 'center',
                                     'valign': 'vcenter',
                                     'fg_color': '#D7E4BC',
                                     'border': 1})

center_format = workbook.add_format({'align': 'center'})

worksheet3.freeze_panes(1, 1)

# Other sheet formatting.
worksheet3.set_column('A:Z', 16)
worksheet3.set_row(0, 20)
worksheet3.set_selection('C3')
worksheet3.write(0, 0, '', header_format)

# Some text to demonstrate scrolling.
for col in range(1, 26):
    worksheet3.write(0, col, 'Scroll down', header_format)

for row in range(1, 50):
    worksheet3.write(row, 0, 'Scroll right', header_format)
    for col in range(1, 26):
        worksheet3.write(row, col, col, center_format)
workbook.close()
